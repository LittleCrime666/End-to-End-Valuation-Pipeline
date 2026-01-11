import os
import sys
import glob
import pandas as pd
import numpy as np
from typing import List, Union, Optional
import warnings
import time
import openpyxl
from concurrent.futures import ThreadPoolExecutor, as_completed

# Suppress openpyxl UserWarning (termasuk Data Validation extension warning)
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ----------------------------
# KONFIGURASI (input setelah run)
# ----------------------------
def get_kurs_usd_to_idr(year, quarter):
    kurs_file = r'C:\Users\ASUS\Documents\Investasi\Kurs USD.xlsx'
    try:
        df_kurs = pd.read_excel(kurs_file)
        # Normalisasi nama kolom
        df_kurs.columns = [str(c).strip().lower() for c in df_kurs.columns]
        tahun_col = next((c for c in df_kurs.columns if 'tahun' in c), None)
        kuartal_col = next((c for c in df_kurs.columns if 'kuartal' in c), None)
        nilai_col = next((c for c in df_kurs.columns if 'nilai' in c or 'kurs' in c), None)

        if not tahun_col or not kuartal_col or not nilai_col:
            print(f"Kolom di file Kurs USD.xlsx harus mengandung 'tahun', 'kuartal', dan 'nilai'/'kurs'. Kolom ditemukan: {df_kurs.columns}")
            sys.exit(1)

        match = df_kurs[
            (df_kurs[tahun_col].astype(int) == int(year)) &
            (df_kurs[kuartal_col].astype(int) == int(quarter))
        ]
        if not match.empty:
            return float(match.iloc[0][nilai_col])
        else:
            print(f"Kurs USD untuk tahun {year} kuartal {quarter} tidak ditemukan di file Kurs USD.xlsx.")
            sys.exit(1)
    except Exception as e:
        print(f"Gagal membaca file Kurs USD.xlsx: {e}")
        sys.exit(1)

def get_user_input():
    try:
        year = int(input("Masukkan tahun laporan (misal 2025): "))
        quarter = int(input("Masukkan kuartal (1-4): "))
        kurs = get_kurs_usd_to_idr(year, quarter)
        print(f"Kurs USD ke IDR untuk {year} Q{quarter}: {kurs}")
        return year, quarter, kurs
    except Exception:
        print("Input tidak valid. Silakan jalankan ulang.")
        sys.exit(1)

YEAR, QUARTER, KURS_USD_TO_IDR = get_user_input()

# Mulai stopwatch setelah input kurs
start_time = time.time()

# Path (edit sekali)
BASE_INPUT_PATH = r'C:\Users\ASUS\Documents\Investasi\Laporan Keuangan'
OUTPUT_FOLDER = r'C:\Users\ASUS\Documents\Investasi\Rekap Analisa Fundamental\ID'
# ----------------------------
# End configuration
# ----------------------------

# Variabel otomatis (jangan diedit)
INPUT_FOLDER = os.path.join(BASE_INPUT_PATH, f"{YEAR} Q{QUARTER}")
OUTPUT_FILENAME = f"{YEAR} Kuartal {QUARTER}.xlsx"

_ANNUALIZATION_MAP = {1: 4, 2: 2, 3: 4/3, 4: 1}
if QUARTER not in _ANNUALIZATION_MAP:
    print(f"Error: QUARTER must be 1-4 (got {QUARTER}). Exiting.")
    sys.exit(1)
ANNUALIZATION_FACTOR = _ANNUALIZATION_MAP[QUARTER]

# ----------------------------
# HELPER FUNCTIONS (Cepat)
# ----------------------------

def parse_number(val) -> Optional[float]:
    """Mengubah nilai sel Excel menjadi float. Mengembalikan None jika tidak bisa."""
    if pd.isna(val):
        return None
    if isinstance(val, (int, float, np.number)):
        return float(val)
    s = str(val).strip()
    if s == '':
        return None
    neg = False
    if s.startswith('(') and s.endswith(')'):
        neg = True
        s = s[1:-1].strip()
    if s.endswith('%'):
        try:
            v = float(s.rstrip('%').replace(',', '').replace(' ', '')) / 100.0
            return -v if neg else v
        except Exception:
            return None
    s = s.replace(',', '').replace(' ', '')
    try:
        v = float(s)
        return -v if neg else v
    except Exception:
        return None

def safe_div(a, b):
    try:
        a = 0.0 if pd.isna(a) else float(a)
        b = 0.0 if pd.isna(b) else float(b)
        if b == 0:
            return 0.0
        return a / b
    except Exception:
        return 0.0

def find_column_ci(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    """Mencari kolom di df dengan nama case-insensitive di antara kandidat."""
    cols_lower = {c.lower(): c for c in df.columns}
    for cand in candidates:
        c = cand.lower()
        if c in cols_lower:
            return cols_lower[c]
    for cand in candidates:
        for col in df.columns:
            if cand.lower() in str(col).lower():
                return col
    return None

def _norm(s: str) -> str:
    """Helper normalisasi string untuk pencarian."""
    return str(s).strip().lower()

def fast_extract_all_metrics(xlsx_path: str) -> dict:
    """
    Ekstraksi super cepat: scan semua sheet 1x, hanya kolom A (label) dan B (nilai)
    menggunakan openpyxl read_only. Mencari SEMUA key yang dibutuhkan.
    """
    # Definisikan semua kunci yang dibutuhkan
    
    # Kunci yang harus cocok PERSIS (exact match)
    exact_map = {
        _norm("Jumlah laba (rugi) sebelum pajak penghasilan"): "Laba Usaha",
        _norm("Jumlah laba (rugi)"): "Laba Bersih",
        _norm("Jumlah liabilitas"): "Total Liabilitas",
        _norm("Jumlah aset"): "Total Aset",
        _norm("Kode entitas"): "Saham",
    }
    
    # Kunci yang boleh substring (contains)
    contains_map = {
        # Metadata
        "mata uang pelaporan": "Mata uang pelaporan",
        "pembulatan yang digunakan": "Pembulatan",
        # Data Keuangan
        "jumlah aset lancar": "Aset Lancar",
        "jumlah aset tidak lancar": "Aset Tetap",
        "jumlah liabilitas jangka pendek": "Liabilitas Jangka Pendek",
        "jumlah liabilitas jangka panjang": "Liabilitas Jangka Panjang",
        "jumlah dana syirkah temporer": "Dana Syirkah Temporer",
        "jumlah ekuitas yang diatribusikan kepada pemilik entitas induk": "Ekuitas",
        "jumlah laba bruto": "Laba Bruto",
        "jumlah arus kas bersih yang diperoleh dari (digunakan untuk) aktivitas operasi": "Arus Kas Operasi",
        "jumlah arus kas bersih yang diperoleh dari (digunakan untuk) aktivitas investasi": "Arus Kas Investasi",
        "jumlah arus kas bersih yang diperoleh dari (digunakan untuk) aktivitas pendanaan": "Arus Kas Pendanaan",
        # Pendapatan (dengan prioritas)
        "penjualan dan pendapatan usaha": "Pendapatan",
        "pendapatan bunga": "Pendapatan",
        "pendapatan dari premi asuransi": "Pendapatan",
    }
    
    # Prioritas khusus untuk Pendapatan
    pendapatan_priority = {
        "penjualan dan pendapatan usaha": 0,
        "pendapatan bunga": 1,
        "pendapatan dari premi asuransi": 2,
    }

    # Kumpulan semua output kolom yang kita cari
    all_keys = {
        "Saham", "Mata uang pelaporan", "Pembulatan", "Aset Lancar", "Aset Tetap",
        "Total Aset", "Liabilitas Jangka Pendek", "Liabilitas Jangka Panjang",
        "Dana Syirkah Temporer", "Total Liabilitas", "Ekuitas", "Pendapatan",
        "Laba Bruto", "Laba Usaha", "Laba Bersih", "Arus Kas Operasi",
        "Arus Kas Investasi", "Arus Kas Pendanaan"
    }
    
    needed = set(all_keys)
    result = {k: None for k in all_keys}
    best_pendapatan = (999, None) # (priority, value)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    except Exception as e:
        # Gagal buka file, kembalikan hasil kosong dengan error
        result['_error'] = f"Gagal buka file: {e}"
        return result

    for ws in wb.worksheets:
        if not needed: # Jika semua sudah ketemu, stop
            break
            
        try:
            # iter_rows(values_only=True) sangat cepat
            for label, value in ws.iter_rows(min_col=1, max_col=2, values_only=True):
                if label is None or value in (None, ""):
                    continue
                
                lbl = _norm(label)
                if lbl == "":
                    continue

                # 1. Cek Exact Match
                if lbl in exact_map:
                    out = exact_map[lbl]
                    if out in needed or result[out] is None:
                        result[out] = value
                        needed.discard(out)

                # 2. Cek Contains Match
                for sub, out in contains_map.items():
                    if sub in lbl:
                        if out == "Pendapatan":
                            pr = pendapatan_priority.get(sub, 99)
                            if pr < best_pendapatan[0]:
                                best_pendapatan = (pr, value)
                                result["Pendapatan"] = value
                                needed.discard("Pendapatan")
                        elif out in needed or result[out] is None:
                            result[out] = value
                            needed.discard(out)
                            
                if not needed: # Early-stop
                    break
        except Exception:
            # Abaikan sheet yang error (misal: sheet terproteksi/aneh)
            continue

    try:
        wb.close()
    except Exception:
        pass
        
    return result

def process_file_worker(xlsx_path: str, kurs_usd: float, existing_saham_set: set) -> dict:
    """
    Worker function untuk ProcessPoolExecutor.
    Membaca 1 file, mem-parsing, dan mengembalikan 1 baris data (dict).
    """
    
    # 1. Ekstrak semua data dalam 1x pemindaian
    raw_data = fast_extract_all_metrics(xlsx_path)
    
    if '_error' in raw_data:
        return {'_error': raw_data['_error'], '_file': xlsx_path}

    # 2. Dapatkan Kode Saham
    saham_val = raw_data.get('Saham')
    saham_str = str(saham_val).strip() if saham_val is not None else None

    if not saham_str:
        return {'_error': 'Kode Saham tidak ditemukan', '_file': xlsx_path}

    # 3. Cek jika saham sudah ada di file output
    if saham_str in existing_saham_set:
        return {'_skipped': True, 'Saham': saham_str}

    # 4. Tentukan Multiplier Mata Uang
    currency_val = raw_data.get('Mata uang pelaporan')
    currency_multiplier = 1.0
    if isinstance(currency_val, str):
        cv = currency_val.lower()
        if "dollar" in cv or "usd" in cv:
            currency_multiplier = kurs_usd

    # 5. Tentukan Multiplier Pembulatan
    rounding_val = raw_data.get('Pembulatan')
    rounding_multiplier = 1.0
    if isinstance(rounding_val, str):
        rv = rounding_val.lower()
        if "ribuan" in rv or "thousand" in rv:
            rounding_multiplier = 1_000.0
        elif "jutaan" in rv or "million" in rv:
            rounding_multiplier = 1_000_000.0
        elif "miliaran" in rv or "billion" in rv:
            rounding_multiplier = 1_000_000_000.0

    total_multiplier = currency_multiplier * rounding_multiplier

    # 6. Siapkan baris hasil, parse, dan kalikan
    row = {'Saham': saham_str}
    
    financial_keys = [
        "Aset Lancar", "Aset Tetap", "Total Aset", "Liabilitas Jangka Pendek",
        "Liabilitas Jangka Panjang", "Dana Syirkah Temporer", "Total Liabilitas",
        "Ekuitas", "Pendapatan", "Laba Bruto", "Laba Usaha", "Laba Bersih",
        "Arus Kas Operasi", "Arus Kas Investasi", "Arus Kas Pendanaan"
    ]
    
    for key in financial_keys:
        raw_val = raw_data.get(key)
        parsed = parse_number(raw_val)
        row[key] = (parsed * total_multiplier) if parsed is not None else 0.0

    row['_source_file'] = os.path.basename(xlsx_path)
    return row

# ----------------------------
# FUNGSI SIMPAN (Robust)
# ----------------------------

def save_workbook_with_autofit(target_path: str, df_data: pd.DataFrame, df_ringkasan: pd.DataFrame, df_rekap: pd.DataFrame) -> str:
    """Menyimpan workbook dan autofit. Otomatis ganti nama jika file terkunci."""
    def _write(path: str):
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            df_data.to_excel(writer, sheet_name='Data', index=False)
            (df_ringkasan if not df_ringkasan.empty else pd.DataFrame()).to_excel(writer, sheet_name='Ringkasan', index=False)
            df_rekap.to_excel(writer, sheet_name='Rekap', index=False)

        # Autofit
        wb = openpyxl.load_workbook(path)
        for ws in wb.worksheets:
            for col in ws.columns:
                max_length = 0
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                for cell in col:
                    val = "" if cell.value is None else str(cell.value)
                    if len(val) > max_length:
                        max_length = len(val)
                ws.column_dimensions[col_letter].width = max_length + 2
        wb.save(path)
        wb.close()

    try:
        _write(target_path)
        return target_path
    except PermissionError:
        base, ext = os.path.splitext(target_path)
        alt_path = f"{base} - new {time.strftime('%Y%m%d-%H%M%S')}{ext}"
        print(f"Target {target_path} sedang dibuka. Menyimpan ke: {alt_path}")
        _write(alt_path)
        return alt_path
    except Exception as e:
        print(f"Gagal menyimpan file: {e}")
        raise

# ----------------------------
# MAIN EXECUTION
# ----------------------------

if __name__ == "__main__":
    if not os.path.isdir(INPUT_FOLDER):
        print(f"Input folder tidak ada: {INPUT_FOLDER}")
        sys.exit(1)

    xlsx_files = [f for f in glob.glob(os.path.join(INPUT_FOLDER, "*.xlsx"))]
    data_files = [f for f in xlsx_files if not os.path.basename(f).startswith("Ringkasan Saham-")]
    ringkasan_files = [f for f in xlsx_files if os.path.basename(f).startswith("Ringkasan Saham-")]

    # 1. Baca file output lama jika ada, untuk mendapatkan daftar saham yang sudah diproses
    existing_saham = set()
    output_path = os.path.join(OUTPUT_FOLDER, OUTPUT_FILENAME)
    if os.path.exists(output_path):
        try:
            old_df = pd.read_excel(output_path, sheet_name='Data')
            if 'Saham' in old_df.columns:
                existing_saham = set(old_df['Saham'].dropna().astype(str))
                print(f"Ditemukan {len(existing_saham)} saham yang sudah ada di {output_path}. Akan dilewati.")
        except Exception as e:
            print(f"Peringatan: Gagal membaca file output lama {output_path}. Error: {e}")
            pass

    # 2. Proses semua file menggunakan ProcessPoolExecutor
    records = []
    files_to_process = data_files
    print(f"Mulai memproses {len(files_to_process)} file (menggunakan parallel processing)...")
    
    start_process_time = time.time()
    
    with ThreadPoolExecutor() as executor:
        # Submit semua pekerjaan
        futures = {executor.submit(process_file_worker, f, KURS_USD_TO_IDR, existing_saham): f for f in files_to_process}
        
        processed_count = 0
        skipped_count = 0
        error_count = 0
        
        for future in as_completed(futures):
            file_path = futures[future]
            processed_count += 1
            
            try:
                result_row = future.result()
                
                if not result_row:
                    continue
                
                if '_skipped' in result_row:
                    skipped_count += 1
                elif '_error' in result_row:
                    error_count += 1
                    print(f"ERROR: Gagal memproses {result_row.get('_file', file_path)}: {result_row['_error']}")
                else:
                    records.append(result_row)
                    
            except Exception as e:
                error_count += 1
                print(f"FATAL ERROR: Gagal total pada worker untuk file {file_path}: {e}")

            # Update progress
            if processed_count % 20 == 0 or processed_count == len(files_to_process):
                print(f"Proses... {processed_count}/{len(files_to_process)} file selesai. "
                      f"(Baru: {len(records)}, Lewat: {skipped_count}, Gagal: {error_count})")

    print(f"Selesai memproses {len(files_to_process)} file dalam {time.time() - start_process_time:.2f} detik.")
    print(f"Hasil: {len(records)} data baru, {skipped_count} dilewati, {error_count} gagal.")

    # 3. Gabungkan data baru ke file output lama
    if not records and not os.path.exists(output_path):
        print("Tidak ada data baru yang diproses dan tidak ada file lama. Keluar.")
        sys.exit(0)

    if not records and os.path.exists(output_path):
        print("Tidak ada data baru yang diproses. Membaca data lama saja...")
        try:
            df_data = pd.read_excel(output_path, sheet_name='Data')
        except Exception as e:
            print(f"Gagal membaca data lama: {e}")
            df_data = pd.DataFrame()
    elif records:
        new_df = pd.DataFrame(records)
        # Bagi data BARU dengan 1 Miliar
        cols_to_divide = [
            'Aset Lancar', 'Aset Tetap', 'Total Aset',
            'Liabilitas Jangka Pendek', 'Liabilitas Jangka Panjang', 'Dana Syirkah Temporer',
            'Total Liabilitas', 'Ekuitas', 'Pendapatan', 'Laba Bruto', 'Laba Usaha',
            'Laba Bersih', 'Arus Kas Operasi', 'Arus Kas Investasi', 'Arus Kas Pendanaan'
        ]
        for col in cols_to_divide:
            if col in new_df.columns:
                new_df[col] = pd.to_numeric(new_df[col], errors='coerce').fillna(0) / 1_000_000_000
        
        # Gabungkan
        if os.path.exists(output_path):
            try:
                old_df = pd.read_excel(output_path, sheet_name='Data')
                df_data = pd.concat([old_df, new_df], ignore_index=True)
            except Exception as e:
                print(f"Peringatan: Gagal menggabung dengan data lama. Menyimpan data baru saja. Error: {e}")
                df_data = new_df
        else:
            df_data = new_df
    else:
        # Seharusnya tidak sampai sini, tapi sebagai penjaga
        df_data = pd.DataFrame()

    # Hapus duplikat berdasarkan 'Saham', ambil yang TERAKHIR (data baru)
    if 'Saham' in df_data.columns:
        df_data = df_data.drop_duplicates(subset=['Saham'], keep='last')

    # Hapus kolom _source_file
    if '_source_file' in df_data.columns:
        df_data = df_data.drop(columns=['_source_file'])

    # 4. Proses Ringkasan (tidak berubah dari script Anda)
    df_ringkasan = pd.DataFrame()
    if ringkasan_files:
        try:
            df_ringkasan = pd.read_excel(ringkasan_files[0], sheet_name=0)
        except Exception as e:
            print(f"Peringatan: Gagal membaca file ringkasan {ringkasan_files[0]}: {e}")

    # 5. Buat sheet Rekap (tidak berubah dari script Anda)
    rekap = df_data[['Saham', 'Total Liabilitas', 'Ekuitas', 'Pendapatan', 'Laba Bruto', 'Laba Bersih']].copy()
    for col in ['Total Liabilitas', 'Ekuitas', 'Pendapatan', 'Laba Bruto', 'Laba Bersih']:
        if col in rekap.columns:
            rekap[col] = rekap[col].astype(float).round(1)

    if not df_ringkasan.empty:
        penutupan_col = find_column_ci(df_ringkasan, ["Penutupan", "Penutupan (Close)", "Close", "Last", "Penutupan Harga"])
        tradable_col = find_column_ci(df_ringkasan, ["Tradable Shares", "Tradable", "Shares", "Free Float", "Tradable Shares (Saham)"])
        kode_col = find_column_ci(df_ringkasan, ["Kode Saham", "Kode", "Saham", "Kode saham"])
        
        if kode_col is None:
            kode_col = find_column_ci(df_ringkasan, ["kode", "saham"])
            
        ringkasan_extract = df_ringkasan.copy()
        if kode_col:
            ringkasan_extract = ringkasan_extract.rename(columns={kode_col: 'Kode Saham'})
        else:
            ringkasan_extract['Kode Saham'] = None

        if penutupan_col:
            ringkasan_extract = ringkasan_extract.rename(columns={penutupan_col: 'Penutupan'})
        else:
            ringkasan_extract['Penutupan'] = 0.0
        if tradable_col:
            ringkasan_extract = ringkasan_extract.rename(columns={tradable_col: 'Tradable Shares'})
        else:
            ringkasan_extract['Tradable Shares'] = 0.0

        merge_left = rekap.rename(columns={'Saham': 'Saham'})
        df_merged = merge_left.merge(ringkasan_extract, how='left', left_on='Saham', right_on='Kode Saham', suffixes=('', '_rk'))
    else:
        df_merged = rekap.copy()
        df_merged['Penutupan'] = 0.0
        df_merged['Tradable Shares'] = 0.0

    df_merged['Harga'] = df_merged.get('Penutupan', 0.0).apply(lambda x: parse_number(x) if not pd.isna(x) else 0.0)
    df_merged['Shares_Raw'] = df_merged.get('Tradable Shares', 0.0).apply(lambda x: parse_number(x) if not pd.isna(x) else 0.0)
    df_merged['Shares'] = df_merged['Shares_Raw'].fillna(0.0).astype(float) / 1_000_000_000  # bagi 1 Miliar
    df_merged['Harga'] = df_merged['Harga'].fillna(0.0).astype(float)

    df_merged['Laba Bersih Tahunan'] = df_merged['Laba Bersih'] * ANNUALIZATION_FACTOR
    df_merged['EPS'] = df_merged.apply(lambda r: safe_div(r['Laba Bersih Tahunan'], r['Shares']), axis=1)
    df_merged['Cap'] = df_merged['Harga'] * df_merged['Shares']
    df_merged['PER (x)'] = df_merged.apply(lambda r: safe_div(r['Harga'], r['EPS']), axis=1)
    df_merged['PBV (x)'] = df_merged.apply(lambda r: safe_div(r['Harga'], safe_div(r['Ekuitas'], r['Shares'])), axis=1)
    df_merged['DER (x)'] = df_merged.apply(lambda r: safe_div(r['Total Liabilitas'], r['Ekuitas']), axis=1)
    df_merged['ROE (%)'] = df_merged.apply(lambda r: safe_div(r['Laba Bersih Tahunan'], r['Ekuitas']) * 100.0, axis=1)
    df_merged['GPM (%)'] = df_merged.apply(lambda r: safe_div(r['Laba Bruto'], r['Pendapatan']) * 100.0, axis=1)
    df_merged['NPM (%)'] = df_merged.apply(lambda r: safe_div(r['Laba Bersih'], r['Pendapatan']) * 100.0, axis=1)

    final_cols = ['Saham', 'Harga', 'Shares', 'Total Liabilitas', 'Ekuitas', 'Pendapatan',
                  'Laba Bruto', 'Laba Bersih', 'EPS', 'Cap', 'PER (x)', 'PBV (x)', 'DER (x)',
                  'ROE (%)', 'GPM (%)', 'NPM (%)']
    for c in final_cols:
        if c not in df_merged.columns:
            df_merged[c] = 0.0 if c != 'Saham' else None

    df_rekap = df_merged[final_cols].copy()

    rekap_cols_to_round = ['Shares', 'EPS', 'Cap', 'PER (x)', 'PBV (x)', 'DER (x)', 'ROE (%)', 'GPM (%)', 'NPM (%)']
    for col in rekap_cols_to_round:
        if col in df_rekap.columns:
            df_rekap[col] = df_rekap[col].astype(float).round(1)

    # 6. Simpan ke Excel
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    
    try:
        saved_path = save_workbook_with_autofit(output_path, df_data, df_ringkasan, df_rekap)
        elapsed = time.time() - start_time
        elapsed_hms = time.strftime("%H:%M:%S", time.gmtime(elapsed))  # format HH:MM:SS
        print(f"\nSukses! Output disimpan ke {saved_path}")
        print(f"Total waktu eksekusi: {elapsed_hms} ({elapsed:.2f} detik)")
    except Exception as e:
        elapsed = time.time() - start_time
        elapsed_hms = time.strftime("%H:%M:%S", time.gmtime(elapsed))  # format HH:MM:SS
        print(f"\nGagal menyimpan output: {e}")
        print(f"Total waktu eksekusi (gagal): {elapsed_hms} ({elapsed:.2f} detik)")

